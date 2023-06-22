from django.shortcuts import render,redirect
# from .helpers import analyzeS1
from django.core.files.storage import FileSystemStorage
from django.http import FileResponse,HttpResponseNotFound
from django.shortcuts import HttpResponse
import tabula
import pandas as pd
import os
from django.conf import settings
from result_app.models import Student
import csv
from openpyxl import Workbook
import time
import win32com.client as win32
import pythoncom
import matplotlib.pyplot as plt
from openpyxl.drawing.image import Image
pythoncom.CoInitialize()

# Create your views here.
def login(request):
    if request.method == 'POST':
        uploaded_file = request.FILES['upload']
        fs = FileSystemStorage()
        fs.save(uploaded_file.name, uploaded_file)
        # process the PDF file here
        return HttpResponse('File uploaded successfully!')
    else:
        return render(request, 'result_app/template_file.html')

def upload_pdf(request):
    if request.method == 'POST':
        global uploaded_file
        uploaded_file = request.FILES['upload']
        global semester
        semester = request.POST['semester']
        print(semester)
        fs = FileSystemStorage()
        fs.save(uploaded_file.name, uploaded_file)
        saved_file_path = fs.save(uploaded_file.name, uploaded_file)
        if semester=='S4' :
            result_path=analyze_s4(saved_file_path)
        if semester=='S1':
            result_path=analyze_s1(saved_file_path)
        if semester=='S3':
            result_path=analyze_s3(saved_file_path)
        if semester=='S2':
            result_path=analyze_s2(saved_file_path)
        if semester=='S5':
            result_path=analyze_s5(saved_file_path)
        # process the PDF file here
        download_url = '/upload/upload.html/download.html/'
        return redirect(download_url)
    else:
        return render(request, 'result_app/upload.html')
    
def download_pdf(request):
    if 'resultexcel' in request.POST:
        direct_url='/upload/upload.html/download.html/resultexcel/'
        download_url = request.build_absolute_uri(direct_url)
        return redirect(download_url)
    elif 'resultpdf' in request.POST:
        direct_url='/upload/upload.html/download.html/resultpdf/'
        download_url = request.build_absolute_uri(direct_url)
        return redirect(download_url)
    else:
        return render(request, 'result_app/download.html')

def download_excel_view(request):
    # Get the current directory
    current_dir = os.getcwd()

    # Define the path to the media directory
    media_dir = os.path.join(current_dir, 'media')
    file_path = os.path.join(media_dir,'converted', 'output4.xlsx')
    if os.path.exists(file_path):
        with open(file_path, 'rb') as file:
            response = HttpResponse(file.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="result.xlsx"'
            return response
    else:
        return HttpResponseNotFound("File not found.")

def download_pdf_view(request):
    # Get the current directory
    current_dir = os.getcwd()

    # Define the path to the media directory
    media_dir = os.path.join(current_dir, 'media')
    file_path = os.path.join(media_dir,'converted', 'output4.pdf')
    if os.path.exists(file_path):
        with open(file_path, 'rb') as file:
            response = HttpResponse(file.read(), content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="result.pdf"'
            return response
    else:
        return HttpResponseNotFound("File not found.")

# def download_result_view(request):
#     file_path = os.path.join(settings.MEDIA_ROOT, 'uploads', 'S2Result4.pdf')
#     if os.path.exists(file_path):
#         with open(file_path, 'rb') as file:
#             response = FileResponse(file)
#             response['Content-Disposition'] = 'attachment; filename="result.pdf"'
#             return response
#     else:
#         return HttpResponseNotFound("File not found.")
    
# def download_result_view(request):
#     file_path = os.path.join(settings.MEDIA_ROOT, 'uploads', 'S2Result4.pdf')
#     if os.path.exists(file_path):
#         with open(file_path, 'rb') as file:
#             response = FileResponse(file)
#             response['Content-Disposition'] = 'attachment; filename="result.pdf"'
#             return response
#     else:
#         return HttpResponseNotFound("File not found.")
def create_empty_csv(directory, filename):
    # Create an empty CSV file in the specified directory
    df = pd.DataFrame()
    file_path = os.path.join(directory, filename)
    df.to_csv(file_path, index=False)
    return file_path


def analyze_s4(request):
    
    # Get the current directory
    current_dir = os.getcwd()

    # Define the path to the media directory
    media_dir = os.path.join(current_dir, 'media')

    # Define the path to the uploads directory inside the media directory
    uploads_dir = os.path.join(media_dir, 'uploads')

    # Create the uploads directory if it doesn't exist
    if not os.path.exists(uploads_dir):
        os.makedirs(uploads_dir)

    # Initialize csv_path variable
    csv_path = None

    try:
        file = tabula.read_pdf(uploaded_file, pages="all")
        tabula.convert_into(uploaded_file, os.path.join(uploads_dir, "output2.csv"), output_format="csv", pages="all")
        csv_path = os.path.join(uploads_dir, "output2.csv")
    except Exception as e:
        print("Error:", e)


    file = file

    li=[]
    
    # Open the CSV file
    with open(csv_path, 'r') as file:
        # Create a CSV reader object
        reader = csv.reader(file)

        # Iterate over each row in the CSV file
        for row in reader:
            # Access the values in each row
            # Example: Print the values in each row
            for i in row:
                if i.startswith('TRV20IT'):
                    li.append((row))
                elif i.startswith('LTRV20IT'):
                    li.append((row))
                elif i.startswith('PKD20IT'):
                    li.append((row))
                else:
                    continue
    time.sleep(1)
    output_file='output.csv'
    create_empty_csv(uploads_dir,output_file)
    length=len(li)
    csv1_path = os.path.join(uploads_dir, "output.csv")
    with open(csv1_path, 'w', newline='') as file1:
        writer = csv.writer(file1)
        for i in range(length):
            writer.writerow(li[i])


    # Function to find index of a substring from a list
    def findsubstring(s1):
        with open(csv1_path, 'r') as file:
            reader=csv.reader(file)
            for s in reader:
                r=s[1].split(",")
                for i, string in enumerate(r):
                    if s1 in string:
                        return i
                return -1 
                
    def print_string_between_parentheses(input_string):
        opening_index = input_string.find('(')
        closing_index = input_string.find(')')
        
        if opening_index != -1 and closing_index != -1:
            substring = input_string[opening_index + 1 : closing_index]
            return substring

    def add_list(st):
        with open(csv1_path, 'r') as file:
            reader = csv.reader(file)
            li = []
            for row in reader:
                r1=row[1].split(",")
                if findsubstring(st) != -1:
                    index = findsubstring(st)

                    li.append(print_string_between_parentheses(r1[index]))
                else:
                    li.append("Absent")
            return li

    def add_regID(filename):
        with open(filename, 'r') as file:
            reader = csv.reader(file)
            li = []
            listTRV = []
            for row in reader:
                if 'TRV' in row[0]:
                    listTRV.append(row[0])
                elif 'PKD' in row[0]:
                    listTRV.append(row[0])
            return listTRV


    grade_points = {'S': 10, 'A+': 9.0, 'A': 8.5, 'B+': 8.0, 'B': 7.5, 'C+': 7.0, 'C': 6.5, 'D': 6.0, 'P': 5.5, 'F': 0, 'FE': 0, 'Ab': 0}
    dict={}
    credits = {'MAT101':4,'PHT100':4,'EST110':3,'EST120':4,'HUN101':0,'PHL120':1,'ESL120':1,'CYT100': 4, 'EST100': 4, 'EST130': 4, 'CYL120': 1, 'ESL130': 1, 'MAT102': 4, 'HUN102': 0, 'EST102': 4,'MAT203':4,'ITT201':4,'ITT203':4,'ITT205':4,'EST200':2,'MCN201':0,'ITL201':2,'ITL203':2,'MAT208':4,'ITT202':4,'ITT204':4,'ITT206':4,'HUT200':2,'MCN202':0,'ITL202':2,'ITL204':2,'ITT301':4,'ITT303':4,'ITT305':4,'ITT307':4,'ITT309':3,'MCN301':0,'ITL331':2,'ITL333':2,'ITT302':4,'ITT304':4,'ITT306':4,'ITT308':1,'ITL332':2,'ITD334':2,'ITT312':3,'ITT332':3}
    r1=[]
    # def get_first_row_as_dict(file_path):
    time.sleep(1)
    with open(csv1_path, 'r') as csvfile:
        reader = csv.reader(csvfile)
        for i in reader:
            r1=i[1].split(",")
            # print(r1)
        for j in r1:
            index=j.index("(")
            # print(index)
            dict[j[index-6:index]]=credits[j[index-6:index]]

    for i in dict:
        list_name="list"+str(i)
        list_name=add_list(str(i))
        # print(list_name)
        length=len(list_name)

    listTRV = add_regID(csv1_path)
    l=list(dict.keys())
    # with open('output4.csv', 'w', newline='') as file1:
    #     writer = csv.writer(file1)
        
    #     writer.writerow(['RegID', l[0], l[1],l[2],l[3],l[4],l[5],l[6],l[7]])
    #     for i in range(length):
    #         writer.writerow([listTRV[i],add_list(l[0])[i],add_list(l[1])[i],add_list(l[2])[i],add_list(l[3])[i],add_list(l[4])[i],
    #                          add_list(l[5])[i],add_list(l[6])[i],add_list(l[7])[i]])
    listc1=add_list(l[0])
    listc2=add_list(l[1])
    listc3=add_list(l[2])
    listc4=add_list(l[3])
    listc5=add_list(l[4])
    listc6=add_list(l[5])
    listc7=add_list(l[6])
    listc8=add_list(l[7])

    listnames=[listc1,listc2,listc3,listc4,listc5,listc6,listc7,listc8]
    # Calculate SGPA
                    
    sgpa=0
    grade_counts = {course_code: {'S': 0, 'A+': 0, 'A': 0, 'B+': 0, 'B': 0, 'C+': 0, 'C': 0, 'D': 0, 'P': 0, 'F': 0, 'FE': 0, 'Absent': 0} for course_code in dict}
    SGPA=[]
    grade_points = {'S': 10, 'A+': 9.0, 'A': 8.5, 'B+': 8.0, 'B': 7.5, 'C+': 7.0, 'C': 6.5, 'D': 6.0, 'P': 5.5, 'F': 0, 'FE': 0, 'Absent': 0}
    credits = {'MAT101':4,'PHT101':4,'EST110':3,'EST120':4,'HUN101':0,'PHL120':1,'ESL120':1,'CYT100': 4, 'EST100': 4, 'EST130': 4, 'CYL120': 1, 'ESL130': 1, 'MAT102': 4, 'HUN102': 0, 'EST102': 4,'MAT203':4,'ITT201':4,'ITT203':4,'ITT205':4,'EST200':2,'MCN201':0,'ITL201':2,'ITL203':2,'MAT208':4,'ITT202':4,'ITT204':4,'ITT206':4,'HUT200':2,'MCN202':0,'ITL202':2,'ITL204':2,'ITT301':4,'ITT303':4,'ITT305':4,'ITT307':4,'ITT309':3,'MCN301':0,'ITL331':2,'ITL333':2,'ITT302':4,'ITT304':4,'ITT306':4,'ITT308':1,'ITL332':2,'ITD334':2,'ITT312':3,'ITT332':3}
    for i in range(length):
        has_failing_grade = False
        total_credit_points = 0
        total_credits = 0
        j=0
        for course_code in dict:
            # grade = eval("list" + course_code)[i]
        
            li=listnames[j]
            # print(li)
            grade=li[i]
            # print(grade,end=" ")
            if grade in grade_points:
                if grade in ['F', 'FE', 'Absent']:
                    grade_counts[course_code][grade] += 1
                    has_failing_grade = True
                else:
                    grade_counts[course_code][grade] += 1
                total_credit_points += credits[course_code] * grade_points[grade]
                total_credits += credits[course_code]
                j=j+1
        if has_failing_grade:
            sgpa = 0
        else:
            sgpa = total_credit_points / total_credits
        SGPA.append(sgpa)
    SGPA1=SGPA

    def find_top_students(filename, sgpa_list, reg_id_list):
        student_data = zip(reg_id_list, sgpa_list)
        sorted_students = sorted(student_data, key=lambda x: x[1], reverse=True)
        top_students = sorted_students[:3]
        return top_students
    
    def update_student(stu_id, sem, sgpa):
     try:
        # Try to retrieve the existing student record
        student = Student.objects.get(stu_id=stu_id, sem=sem)
        
        # If a matching record is found, update the existing row
        student.sgpa = sgpa
        student.save()
     except Student.DoesNotExist:
        # If the combination doesn't exist, create a new row
        student = Student(stu_id=stu_id, sem=sem, sgpa=sgpa)
        student.save()

        
    # Write the result to a new CSV file
    time.sleep(1)
    output_file='output4.csv'
    create_empty_csv(uploads_dir,output_file)
    csv4_path = os.path.join(uploads_dir, "output4.csv")
    with open(csv4_path, 'w', newline='') as file1:
        writer = csv.writer(file1)
        writer.writerow(['RegID', l[0], l[1],l[2],l[3],l[4],l[5],l[6],l[7],'SGPA'])
        for i in range(length):
           update_student(listTRV[i],'S4',SGPA[i])

        # Assign the new values to the corresponding fields of the student object
                    
           writer.writerow([listTRV[i],add_list(l[0])[i],add_list(l[1])[i],add_list(l[2])[i],add_list(l[3])[i],add_list(l[4])[i],
                            add_list(l[5])[i],add_list(l[6])[i],add_list(l[7])[i],SGPA[i]])

    #         if semester=="S5":
    #             if check_combination_exists(listTRV[i],"S5"):
    #                 new_stu_id = listTRV[i]
    #                 new_sem = "S5"
    #                 new_sgpa =SGPA[i]

    #     # Assign the new values to the corresponding fields of the student object
    #                 student.stu_id = new_stu_id
    #                 student.sem = new_sem
    #                 student.sgpa = new_sgpa

    # # Perform a bulk update to update all the students in a single query
    #                 Student.objects.bulk_update(students, ['stu_id', 'sem', 'sgpa'])
    #             else:
    #              student = Student(stu_id=listTRV[i], sem="S5", sgpa=SGPA[i])
    #              student.save()

            
            
    # Iterate over each student and update their values
    
        # Calculate new values for stu_id, sem, and sgpa based on your logic
        
                


    top_students=find_top_students(csv1_path, SGPA, listTRV)
    # Convert the CSV to Excel format
    wb = Workbook()
    ws = wb.active

    time.sleep(1)
    with open(csv4_path, 'r') as file2:
        reader = csv.reader(file2)
        for row in reader:
            ws.append(row)
    ws.append([])  # Add an empty row for separation
    ws.append(["Top Three Students"])
    ws.append(["RegID", "SGPA"])
    for student in top_students:
        ws.append(student)
    # Add grade counts per subject to the Excel file
    ws.append([])  # Add an empty row for separation
    ws.append(["Grade Counts per Subject"])
    for course_code, grade_count in grade_counts.items():
        ws.append([])  # Add an empty row for separation
        ws.append(["Subject: " + course_code])
        ws.append(["Grade", "Count"])
        for grade, count in grade_count.items():
            ws.append([grade, count])
    ws.append([])  # Add an empty row for separation
    ws.append(["Supplementary Count for Each Course"])
    ws.append(['Course', 'Absent', 'F', 'FE'])
    # for course_code in dict:
    #     ws.append([course_code, (grade_count.get(course_code, 0)), count_F.get(course_code, 0), count_FE.get(course_code, 0)])
    for course_code,grades in grade_counts.items():
        fe_count=grades.get('FE',0)
        f_count=grades.get('F',0)
        ab_count=grades.get('Absent',0)
        ws.append([course_code,ab_count, f_count, fe_count])

    dict_passfail={course_code: {'passcount': 0, 'failcount': 0} for course_code in dict}
    for i in range(length):
        j=0
        for course_code in dict:
            # grade = eval("list" + course_code)[i]
        
            li=listnames[j]
            # print(li)
            grade=li[i]
            # print(grade,end=" ")
            if grade in grade_points:
                if grade in ['F', 'FE', 'Absent']:
                    dict_passfail[course_code]['failcount'] += 1
                else:
                    dict_passfail[course_code]['passcount'] += 1
                j=j+1

    dict_passpercent={course_code: 0 for course_code in dict}
    percent=0
    list_codes=list(dict.keys())
    j=0
    for i in dict_passfail.values():
            course_code=list_codes[j]
            # print(i['passcount'])
            percent=i['passcount']*100/(i['passcount']+i['failcount'])
            dict_passpercent[course_code]=percent
            j=j+1
            
            # print(percent)

    # print(dict_passfail)
    ws.append([])  # Add an empty row for separation
    ws.append(["Pass Percentage for Each Course"])
    ws.append(['Course','Percent'])
    # for course_code in dict:
    #     ws.append([course_code, (grade_count.get(course_code, 0)), count_F.get(course_code, 0), count_FE.get(course_code, 0)])
    for course_code,percent in dict_passpercent.items():
            ws.append([course_code,percent])
            plt.bar(course_code,percent)
            plt.xlabel('Course')
            plt.ylabel('Passpercent')
            plt.title('Bar Chart')

# Save the figure
            plt.savefig('chart.png')
            img = Image('chart.png')
            img.width = 400
            img.height = 300
            ws.add_image(img, 'E213')
    CGPA=[]
    cgpa=0
    for i in range(length):
     try:
         student1=Student.objects.get(stu_id=listTRV[i], sem='S1')
         student2=Student.objects.get(stu_id=listTRV[i], sem='S2')
         student3=Student.objects.get(stu_id=listTRV[i], sem='S3')
         student4=Student.objects.get(stu_id=listTRV[i], sem='S4')
         if student1.sgpa==0 or student2.sgpa==0 or student3.sgpa==0 or student4.sgpa==0:
             cgpa=0
         else:
          cgpa=((student1.sgpa*17)+(student2.sgpa*21)+(student3.sgpa*22)+(student4.sgpa*22))/82
         CGPA.append(cgpa)
     except Student.DoesNotExist:
         student3=Student.objects.get(stu_id=listTRV[i], sem='S3')
         student4=Student.objects.get(stu_id=listTRV[i], sem='S4')
         if student3.sgpa==0 or student4.sgpa==0:
             cgpa=0
         else:
          cgpa=((student3.sgpa*22)+(student4.sgpa*22))/44
         CGPA.append(cgpa)
    ws.append([])
    ws.append(["CGPA Calculation"])
    ws.append(['RegID','CGPA'])
    for i in range(length):
        ws.append([listTRV[i],CGPA[i]])
    # wb.save('output4.xlsx')


    output_path = os.path.join("media", "converted", "output4.xlsx")
    # workbook = Workbook()
    wb.save(output_path)

    pythoncom.CoInitialize()
    # Open the Excel application
    excel = win32.gencache.EnsureDispatch('Excel.Application')

    # Open the workbook
    wb = excel.Workbooks.Open(r'C:\Users\Reon Augustine\Desktop\MiniProject\virtualresult\result_analysis\media\converted\output4.xlsx')

    # Save the workbook as a PDF file
    pdf_file_path = r'C:\Users\Reon Augustine\Desktop\MiniProject\virtualresult\result_analysis\media\converted\output4.pdf'
    xlFixedFormat = 0

    try:
        wb.ExportAsFixedFormat(xlFixedFormat, pdf_file_path)
        print(f"PDF file saved successfully: {pdf_file_path}")
    except Exception as e:
        print(f"Error occurred while saving PDF: {e}")

    # Close the workbook
    wb.Close()

    # Quit Excel application
    excel.Quit()

    return output_path







def analyze_s1(request):
     # Get the current directory
    current_dir = os.getcwd()

    # Define the path to the media directory
    media_dir = os.path.join(current_dir, 'media')

    # Define the path to the uploads directory inside the media directory
    uploads_dir = os.path.join(media_dir, 'uploads')

    # Create the uploads directory if it doesn't exist
    if not os.path.exists(uploads_dir):
        os.makedirs(uploads_dir)

    # Initialize csv_path variable
    csv_path = None

    try:
        file = tabula.read_pdf(uploaded_file, pages="all")
        tabula.convert_into(uploaded_file, os.path.join(uploads_dir, "output2.csv"), output_format="csv", pages="all")
        csv_path = os.path.join(uploads_dir, "output2.csv")
    except Exception as e:
        print("Error:", e)


    file = file

    li=[]
    # Open the CSV file
    with open(csv_path, 'r') as file:
        # Create a CSV reader object
        reader = csv.reader(file)

        # Iterate over each row in the CSV file
        for row in reader:
            # Access the values in each row
            # Example: Print the values in each row
            for i in row:
                if i.startswith('TRV20IT'):
                    li.append((row))
                elif i.startswith('LTRV20IT'):
                    li.append((row))
                elif i.startswith('PKD20IT'):
                    li.append((row))
                else:
                    continue
    time.sleep(1)
    output_file='output.csv'
    create_empty_csv(uploads_dir,output_file)
    length=len(li)
    csv1_path = os.path.join(uploads_dir, "output.csv")
    with open(csv1_path, 'w', newline='') as file1:
        writer = csv.writer(file1)
        for i in range(length):
            writer.writerow(li[i])


    # Function to find index of a substring from a list
    def findsubstring(s1):
         with open(csv1_path, 'r') as file:
            reader=csv.reader(file)
            for s in reader:
                r=s[1].split(",")
                l2=list(r)
                for j in l2:
                 if s1 in j:
                    return(l2.index(j))
                    break
            else:
                return(-1)
                
    def print_string_between_parentheses(input_string):
        opening_index = input_string.find('(')
        closing_index = input_string.find(')')
        
        if opening_index != -1 and closing_index != -1:
            substring = input_string[opening_index + 1 : closing_index]
            return substring

    def add_list(st):
        with open(csv1_path, 'r') as file:
            reader = csv.reader(file)
            li = []
            for row in reader:
                r1=row[1].split(",")
                if findsubstring(st) != -1:
                    index = findsubstring(st)

                    li.append(print_string_between_parentheses(r1[index]))
                else:
                    li.append("Absent")
            return li

    def add_regID(filename):
        with open(filename, 'r') as file:
            reader = csv.reader(file)
            li = []
            listTRV = []
            for row in reader:
                if 'TRV' in row[0]:
                    listTRV.append(row[0])
                elif 'PKD' in row[0]:
                    listTRV.append(row[0])
            return listTRV


    grade_points = {'S': 10, 'A+': 9.0, 'A': 8.5, 'B+': 8.0, 'B': 7.5, 'C+': 7.0, 'C': 6.5, 'D': 6.0, 'P': 5.5, 'F': 0, 'FE': 0, 'Ab': 0}
    credits = {'MAT101':4,'PHT100':4,'EST110':3,'EST120':4,'HUN101':0,'PHL120':1,'ESL120':1}
    r1=[]
   
    time.sleep(1)
    listTRV = add_regID(csv1_path)
    l=list(credits.keys())
    listc1=add_list(l[0])
    listc2=add_list(l[1])
    listc3=add_list(l[2])
    listc4=add_list(l[3])
    listc5=add_list(l[4])
    listc6=add_list(l[5])
    listc7=add_list(l[6])
    
                           

    listnames=[listc1,listc2,listc3,listc4,listc5,listc6,listc7]
    # Calculate SGPA
                    
    sgpa=0
    grade_counts = {course_code: {'S': 0, 'A+': 0, 'A': 0, 'B+': 0, 'B': 0, 'C+': 0, 'C': 0, 'D': 0, 'P': 0, 'F': 0, 'FE': 0, 'Absent': 0} for course_code in credits}
    SGPA=[]
    grade_points = {'S': 10, 'A+': 9.0, 'A': 8.5, 'B+': 8.0, 'B': 7.5, 'C+': 7.0, 'C': 6.5, 'D': 6.0, 'P': 5.5, 'F': 0, 'FE': 0, 'Absent': 0}
    for i in range(length):
        has_failing_grade = False
        total_credit_points = 0
        total_credits = 0
        j=0
        for course_code in credits:
            # grade = eval("list" + course_code)[i]
        
            li=listnames[j]
            # print(li)
            grade=li[i]
            # print(grade,end=" ")
            if grade in grade_points:
                if grade in ['F', 'FE', 'Absent']:
                    grade_counts[course_code][grade] += 1
                    has_failing_grade = True
                else:
                    grade_counts[course_code][grade] += 1
                total_credit_points += credits[course_code] * grade_points[grade]
                total_credits += credits[course_code]
                j=j+1
        if has_failing_grade:
            sgpa = 0
        else:
            sgpa = total_credit_points / 17
        SGPA.append(sgpa)

    def find_top_students(filename, sgpa_list, reg_id_list):
        student_data = zip(reg_id_list, sgpa_list)
        sorted_students = sorted(student_data, key=lambda x: x[1], reverse=True)
        top_students = sorted_students[:3]
        return top_students
        

    # function to check if a particular combination of stu and sem already exists in the table
    # def check_combination_exists(stu_id, sem):
    #     combination_exists = Student.objects.filter(stu_id=stu_id, sem=sem).exists()
    #     return combination_exists
    # # to insert new rows to the table
    # def insert_student(stu_id, sem, sgpa):
    #                 student = Student(stu_id=stu_id, sem=sem, sgpa=sgpa)
    #                 student.save()
    # Write the result to a new CSV file
    def update_student(stu_id, sem, sgpa):
     try:
        # Try to retrieve the existing student record
        student = Student.objects.get(stu_id=stu_id, sem=sem)
        
        # If a matching record is found, update the existing row
        student.sgpa = sgpa
        student.save()
     except Student.DoesNotExist:
        # If the combination doesn't exist, create a new row
        student = Student(stu_id=stu_id, sem=sem, sgpa=sgpa)
        student.save()
    time.sleep(1)
    output_file='output4.csv'
    create_empty_csv(uploads_dir,output_file)
    csv4_path = os.path.join(uploads_dir, "output4.csv")
    with open(csv4_path, 'w', newline='') as file1:
        writer = csv.writer(file1)
        writer.writerow(['RegID', l[0], l[1],l[2],l[3],l[4],l[5],l[6],'SGPA'])
        for i in range(length):
            update_student(listTRV[i],'S1',SGPA[i])
            writer.writerow([listTRV[i],add_list(l[0])[i],add_list(l[1])[i],add_list(l[2])[i],add_list(l[3])[i],add_list(l[4])[i],
                            add_list(l[5])[i],add_list(l[6])[i],SGPA[i]])#write cgpa

            #save data add id,add sem,add sgpa
            #get all sgpas of listTRV[i]
            #cgpa calculation
            

# save in this loop^^(database)

    top_students=find_top_students(csv1_path, SGPA, listTRV)
    # Convert the CSV to Excel format
    wb = Workbook()
    ws = wb.active

    time.sleep(1)
    with open(csv4_path, 'r') as file2:
        reader = csv.reader(file2)
        for row in reader:
            ws.append(row)
    ws.append([])  # Add an empty row for separation
    ws.append(["Top Three Students"])
    ws.append(["RegID", "SGPA"])
    for student in top_students:
        ws.append(student)
    # Add grade counts per subject to the Excel file
    ws.append([])  # Add an empty row for separation
    ws.append(["Grade Counts per Subject"])
    for course_code, grade_count in grade_counts.items():
        ws.append([])  # Add an empty row for separation
        ws.append(["Subject: " + course_code])
        ws.append(["Grade", "Count"])
        for grade, count in grade_count.items():
            ws.append([grade, count])
    ws.append([])  # Add an empty row for separation
    ws.append(["Supplementary Count for Each Course"])
    ws.append(['Course', 'Absent', 'F', 'FE'])
    # for course_code in dict:
    #     ws.append([course_code, (grade_count.get(course_code, 0)), count_F.get(course_code, 0), count_FE.get(course_code, 0)])
    for course_code,grades in grade_counts.items():
        fe_count=grades.get('FE',0)
        f_count=grades.get('F',0)
        ab_count=grades.get('Absent',0)
        ws.append([course_code,ab_count, f_count, fe_count])
       

    dict_passfail={course_code: {'passcount': 0, 'failcount': 0} for course_code in credits}
    for i in range(length):
        j=0
        for course_code in credits:
            # grade = eval("list" + course_code)[i]
        
            li=listnames[j]
            # print(li)
            grade=li[i]
            # print(grade,end=" ")
            if grade in grade_points:
                if grade in ['F', 'FE', 'Absent']:
                    dict_passfail[course_code]['failcount'] += 1
                else:
                    dict_passfail[course_code]['passcount'] += 1
                j=j+1

    dict_passpercent={course_code: 0 for course_code in credits}
    percent=0
    list_codes=list(credits.keys())
    j=0
    for i in dict_passfail.values():
            course_code=list_codes[j]
            # print(i['passcount'])
            percent=i['passcount']*100/(i['passcount']+i['failcount'])
            dict_passpercent[course_code]=percent
            j=j+1
            
            # print(percent)

    # print(dict_passfail)
    ws.append([])  # Add an empty row for separation
    ws.append(["Pass Percentage for Each Course"])
    ws.append(['Course','Percent'])
    # for course_code in dict:
    #     ws.append([course_code, (grade_count.get(course_code, 0)), count_F.get(course_code, 0), count_FE.get(course_code, 0)])
    for course_code,percent in dict_passpercent.items():
            ws.append([course_code,percent])
            plt.bar(course_code,percent)
            plt.xlabel('Course')
            plt.ylabel('Passpercent')
            plt.title('Bar Chart')

# Save the figure
            plt.savefig('chart.png')
            img = Image('chart.png')
            img.width = 400
            img.height = 300
            ws.add_image(img, 'E192')
    CGPA=[]
    for i in range(length):
     student = Student.objects.get(stu_id=listTRV[i], sem='S1')
     CGPA.append(student.sgpa)
    ws.append([])
    ws.append(["CGPA Calculation"])
    ws.append(['RegID','CGPA'])
    for i in range(length):
        ws.append([listTRV[i],CGPA[i]])


    output_path = os.path.join("media", "converted", "output4.xlsx")
    # workbook = Workbook()
    wb.save(output_path)

    pythoncom.CoInitialize()
    # Open the Excel application
    excel = win32.gencache.EnsureDispatch('Excel.Application')

    # Open the workbook
    wb = excel.Workbooks.Open(r'C:\Users\Reon Augustine\Desktop\MiniProject\virtualresult\result_analysis\media\converted\output4.xlsx')

    # Save the workbook as a PDF file
    pdf_file_path = r'C:\Users\Reon Augustine\Desktop\MiniProject\virtualresult\result_analysis\media\converted\output4.pdf'
    xlFixedFormat = 0

    try:
        wb.ExportAsFixedFormat(xlFixedFormat, pdf_file_path)
        print(f"PDF file saved successfully: {pdf_file_path}")
    except Exception as e:
        print(f"Error occurred while saving PDF: {e}")

    # Close the workbook
    wb.Close()

    # Quit Excel application
    excel.Quit()

    return output_path




def analyze_s3(request):
     # Get the current directory
    current_dir = os.getcwd()

    # Define the path to the media directory
    media_dir = os.path.join(current_dir, 'media')

    # Define the path to the uploads directory inside the media directory
    uploads_dir = os.path.join(media_dir, 'uploads')

    # Create the uploads directory if it doesn't exist
    if not os.path.exists(uploads_dir):
        os.makedirs(uploads_dir)

    # Initialize csv_path variable
    csv_path = None

    try:
        file = tabula.read_pdf(uploaded_file, pages="all")
        tabula.convert_into(uploaded_file, os.path.join(uploads_dir, "output2.csv"), output_format="csv", pages="all")
        csv_path = os.path.join(uploads_dir, "output2.csv")
    except Exception as e:
        print("Error:", e)


    file = file

    li=[]
    # Open the CSV file
    with open(csv_path, 'r') as file:
        # Create a CSV reader object
        reader = csv.reader(file)

        # Iterate over each row in the CSV file
        for row in reader:
            # Access the values in each row
            # Example: Print the values in each row
            for i in row:
                if i.startswith('TRV20IT'):
                    li.append((row))
                elif i.startswith('LTRV20IT'):
                    li.append((row))
                elif i.startswith('PKD20IT'):
                    li.append((row))
                else:
                    continue
    time.sleep(1)
    output_file='output.csv'
    create_empty_csv(uploads_dir,output_file)
    length=len(li)
    csv1_path = os.path.join(uploads_dir, "output.csv")
    with open(csv1_path, 'w', newline='') as file1:
        writer = csv.writer(file1)
        for i in range(length):
            writer.writerow(li[i])


    # Function to find index of a substring from a list
    def findsubstring(s1):
         with open(csv1_path, 'r') as file:
            reader=csv.reader(file)
            for s in reader:
                r=s[1].split(",")
                l2=list(r)
                for j in l2:
                 if s1 in j:
                    return(l2.index(j))
                    break
            else:
                return(-1)
                
    def print_string_between_parentheses(input_string):
        opening_index = input_string.find('(')
        closing_index = input_string.find(')')
        
        if opening_index != -1 and closing_index != -1:
            substring = input_string[opening_index + 1 : closing_index]
            return substring

    def add_list(st):
        with open(csv1_path, 'r') as file:
            reader = csv.reader(file)
            li = []
            for row in reader:
                r1=row[1].split(",")
                if findsubstring(st) != -1:
                    index = findsubstring(st)

                    li.append(print_string_between_parentheses(r1[index]))
                else:
                    li.append("Absent")
            return li

    def add_regID(filename):
        with open(filename, 'r') as file:
            reader = csv.reader(file)
            li = []
            listTRV = []
            for row in reader:
                if 'TRV' in row[0]:
                    listTRV.append(row[0])
                elif 'PKD' in row[0]:
                    listTRV.append(row[0])
            return listTRV


    grade_points = {'S': 10, 'A+': 9.0, 'A': 8.5, 'B+': 8.0, 'B': 7.5, 'C+': 7.0, 'C': 6.5, 'D': 6.0, 'P': 5.5, 'F': 0, 'FE': 0, 'Ab': 0}
    credits = {'EST200':2,'MCN201':0,'ITT201':4,'ITT203':4,'ITT205':4,'ITL201':2,'ITL203':2,'MAT203':4}
    r1=[]
   
    time.sleep(1)
    listTRV = add_regID(csv1_path)
    l=list(credits.keys())
    listc1=add_list(l[0])
    listc2=add_list(l[1])
    listc3=add_list(l[2])
    listc4=add_list(l[3])
    listc5=add_list(l[4])
    listc6=add_list(l[5])
    listc7=add_list(l[6])
    listc8=add_list(l[7])
    with open(csv1_path, 'r') as csvfile:
        reader = csv.reader(csvfile)
        for row in reader:
            for item in row:
                if "LTRV20IT069"==item :
                    h=next(reader)
                    listc8[0]=print_string_between_parentheses(h[1])


    listnames=[listc1,listc2,listc3,listc4,listc5,listc6,listc7,listc8]
    # Calculate SGPA
                    
    sgpa=0
    grade_counts = {course_code: {'S': 0, 'A+': 0, 'A': 0, 'B+': 0, 'B': 0, 'C+': 0, 'C': 0, 'D': 0, 'P': 0, 'F': 0, 'FE': 0, 'Absent': 0} for course_code in credits}
    SGPA=[]
    grade_points = {'S': 10, 'A+': 9.0, 'A': 8.5, 'B+': 8.0, 'B': 7.5, 'C+': 7.0, 'C': 6.5, 'D': 6.0, 'P': 5.5, 'F': 0, 'FE': 0, 'Absent': 0}
    for i in range(length):
        has_failing_grade = False
        total_credit_points = 0
        total_credits = 0
        j=0
        for course_code in credits:
            # grade = eval("list" + course_code)[i]
        
            li=listnames[j]
            # print(li)
            grade=li[i]
            # print(grade,end=" ")
            if grade in grade_points:
                if grade in ['F', 'FE', 'Absent']:
                    grade_counts[course_code][grade] += 1
                    has_failing_grade = True
                else:
                    grade_counts[course_code][grade] += 1
                total_credit_points += credits[course_code] * grade_points[grade]
                total_credits += credits[course_code]
                j=j+1
        if has_failing_grade:
            sgpa = 0
        else:
            sgpa = total_credit_points / 22
        SGPA.append(sgpa)

    def find_top_students(filename, sgpa_list, reg_id_list):
        student_data = zip(reg_id_list, sgpa_list)
        sorted_students = sorted(student_data, key=lambda x: x[1], reverse=True)
        top_students = sorted_students[:3]
        return top_students
        

    # function to check if a particular combination of stu and sem already exists in the table
    # def check_combination_exists(stu_id, sem):
    #     combination_exists = Student.objects.filter(stu_id=stu_id, sem=sem).exists()
    #     return combination_exists
    # # to insert new rows to the table
    # def insert_student(stu_id, sem, sgpa):
    #                 student = Student(stu_id=stu_id, sem=sem, sgpa=sgpa)
    #                 student.save()
    # Write the result to a new CSV file
    time.sleep(1)
    output_file='output4.csv'
    create_empty_csv(uploads_dir,output_file)
    csv4_path = os.path.join(uploads_dir, "output4.csv")
    def update_student(stu_id, sem, sgpa):
     try:
        # Try to retrieve the existing student record
        student = Student.objects.get(stu_id=stu_id, sem=sem)
        
        # If a matching record is found, update the existing row
        student.sgpa = sgpa
        student.save()
     except Student.DoesNotExist:
        # If the combination doesn't exist, create a new row
        student = Student(stu_id=stu_id, sem=sem, sgpa=sgpa)
        student.save()
    
    with open(csv4_path, 'w', newline='') as file1:
        writer = csv.writer(file1)
        writer.writerow(['RegID', l[0], l[1],l[2],l[3],l[4],l[5],l[6],l[7],'SGPA'])
        for i in range(length):
            
             update_student(listTRV[i],'S3',SGPA[i])
            # students = Student.objects.all()
            # if check_combination_exists(listTRV[i],"S1"):
            #    for student in students: 
            #     new_id=listTRV[i]
            #     new_sem="S1"
            #     new_sgpa=SGPA[i]
            #     student.stu_id = new_id
            #     student.sem = new_sem
            #     student.sgpa = new_sgpa
            #    Student.objects.bulk_update(students, ['stu_id', 'sem', 'sgpa'])
            # else:
            #     insert_student(listTRV[i],"S1",SGPA[i])



            #save data add id,add sem,add sgpa
            #get all sgpas of listTRV[i]
            #cgpa calculation
            
             writer.writerow([listTRV[i],add_list(l[0])[i],add_list(l[1])[i],add_list(l[2])[i],add_list(l[3])[i],add_list(l[4])[i],
                            add_list(l[5])[i],add_list(l[6])[i],add_list(l[7])[i],SGPA[i]])#write cgpa

# save in this loop^^(database)

    top_students=find_top_students(csv1_path, SGPA, listTRV)
    # Convert the CSV to Excel format
    wb = Workbook()
    ws = wb.active

    time.sleep(1)
    with open(csv4_path, 'r') as file2:
        reader = csv.reader(file2)
        for row in reader:
            ws.append(row)
    ws.append([])  # Add an empty row for separation
    ws.append(["Top Three Students"])
    ws.append(["RegID", "SGPA"])
    for student in top_students:
        ws.append(student)
    # Add grade counts per subject to the Excel file
    ws.append([])  # Add an empty row for separation
    ws.append(["Grade Counts per Subject"])
    for course_code, grade_count in grade_counts.items():
        ws.append([])  # Add an empty row for separation
        ws.append(["Subject: " + course_code])
        ws.append(["Grade", "Count"])
        for grade, count in grade_count.items():
            ws.append([grade, count])
    ws.append([])  # Add an empty row for separation
    ws.append(["Supplementary Count for Each Course"])
    ws.append(['Course', 'Absent', 'F', 'FE'])
    # for course_code in dict:
    #     ws.append([course_code, (grade_count.get(course_code, 0)), count_F.get(course_code, 0), count_FE.get(course_code, 0)])
    for course_code,grades in grade_counts.items():
        fe_count=grades.get('FE',0)
        f_count=grades.get('F',0)
        ab_count=grades.get('Absent',0)
        ws.append([course_code,ab_count, f_count, fe_count])
       

    dict_passfail={course_code: {'passcount': 0, 'failcount': 0} for course_code in credits}
    for i in range(length):
        j=0
        for course_code in credits:
            # grade = eval("list" + course_code)[i]
        
            li=listnames[j]
            # print(li)
            grade=li[i]
            # print(grade,end=" ")
            if grade in grade_points:
                if grade in ['F', 'FE', 'Absent']:
                    dict_passfail[course_code]['failcount'] += 1
                else:
                    dict_passfail[course_code]['passcount'] += 1
                j=j+1

    dict_passpercent={course_code: 0 for course_code in credits}
    percent=0
    list_codes=list(credits.keys())
    j=0
    for i in dict_passfail.values():
            course_code=list_codes[j]
            # print(i['passcount'])
            percent=i['passcount']*100/(i['passcount']+i['failcount'])
            dict_passpercent[course_code]=percent
            j=j+1
            
            # print(percent)

    # print(dict_passfail)
    ws.append([])  # Add an empty row for separation
    ws.append(["Pass Percentage for Each Course"])
    ws.append(['Course','Percent'])
    # for course_code in dict:
    #     ws.append([course_code, (grade_count.get(course_code, 0)), count_F.get(course_code, 0), count_FE.get(course_code, 0)])
    for course_code,percent in dict_passpercent.items():
            ws.append([course_code,percent])
            plt.bar(course_code,percent)
            plt.xlabel('Course')
            plt.ylabel('Passpercent')
            plt.title('Bar Chart')

# Save the figure
            plt.savefig('chart.png')
            img = Image('chart.png')
            img.width = 400
            img.height = 300
            ws.add_image(img, 'E215')
    # list_regular=[]
    # list_lateral=[]
    # for i in listTRV:
    #     if i.startswith("TRV20IT"):
    #         list_regular.append(i)
    #     else:
    #         list_lateral.append(i)
    CGPA=[]
    cgpa=0
    for i in range(length):
     try:
         student1=Student.objects.get(stu_id=listTRV[i], sem='S1')
         student2=Student.objects.get(stu_id=listTRV[i], sem='S2')
         student3=Student.objects.get(stu_id=listTRV[i], sem='S3')
         if student1.sgpa==0 or student2.sgpa==0 or student3.sgpa==0:
             cgpa=0
         else:
          cgpa=((student1.sgpa*17)+(student2.sgpa*21)+(student3.sgpa*22))/60
         CGPA.append(cgpa)
     except Student.DoesNotExist:
         student3=Student.objects.get(stu_id=listTRV[i], sem='S3')
         cgpa=student3.sgpa
         CGPA.append(cgpa)
    ws.append([])
    ws.append(["CGPA Calculation"])
    ws.append(['RegID','CGPA'])
    for i in range(length):
        ws.append([listTRV[i],CGPA[i]])
            
    output_path = os.path.join("media", "converted", "output4.xlsx")
    # workbook = Workbook()
    wb.save(output_path)

    pythoncom.CoInitialize()
    # Open the Excel application
    excel = win32.gencache.EnsureDispatch('Excel.Application')

    # Open the workbook
    wb = excel.Workbooks.Open(r'C:\Users\Reon Augustine\Desktop\MiniProject\virtualresult\result_analysis\media\converted\output4.xlsx')

    # Save the workbook as a PDF file
    pdf_file_path = r'C:\Users\Reon Augustine\Desktop\MiniProject\virtualresult\result_analysis\media\converted\output4.pdf'
    xlFixedFormat = 0

    try:
        wb.ExportAsFixedFormat(xlFixedFormat, pdf_file_path)
        print(f"PDF file saved successfully: {pdf_file_path}")
    except Exception as e:
        print(f"Error occurred while saving PDF: {e}")

    # Close the workbook
    wb.Close()

    # Quit Excel application
    excel.Quit()

    return output_path



def analyze_s2(request):
     # Get the current directory
    current_dir = os.getcwd()

    # Define the path to the media directory
    media_dir = os.path.join(current_dir, 'media')

    # Define the path to the uploads directory inside the media directory
    uploads_dir = os.path.join(media_dir, 'uploads')

    # Create the uploads directory if it doesn't exist
    if not os.path.exists(uploads_dir):
        os.makedirs(uploads_dir)

    # Initialize csv_path variable
    csv_path = None

    try:
        file = tabula.read_pdf(uploaded_file, pages="all")
        tabula.convert_into(uploaded_file, os.path.join(uploads_dir, "output2.csv"), output_format="csv", pages="all")
        csv_path = os.path.join(uploads_dir, "output2.csv")
    except Exception as e:
        print("Error:", e)


    file = file

    li=[]
    # Open the CSV file
    with open(csv_path, 'r') as file:
        # Create a CSV reader object
        reader = csv.reader(file)

        # Iterate over each row in the CSV file
        for row in reader:
            # Access the values in each row
            # Example: Print the values in each row
            for i in row:
                if i.startswith('TRV20IT'):
                    li.append((row))
                elif i.startswith('LTRV20IT'):
                    li.append((row))
                elif i.startswith('PKD20IT'):
                    li.append((row))
                else:
                    continue
    time.sleep(1)
    output_file='output.csv'
    create_empty_csv(uploads_dir,output_file)
    length=len(li)
    csv1_path = os.path.join(uploads_dir, "output.csv")
    with open(csv1_path, 'w', newline='') as file1:
        writer = csv.writer(file1)
        for i in range(length):
            writer.writerow(li[i])


    # Function to find index of a substring from a list
    def findsubstring(s1):
         with open(csv1_path, 'r') as file:
            reader=csv.reader(file)
            for s in reader:
                r=s[1].split(",")
                l2=list(r)
                for j in l2:
                 if s1 in j:
                    return(l2.index(j))
                    break
            else:
                return(-1)
                
    def print_string_between_parentheses(input_string):
        opening_index = input_string.find('(')
        closing_index = input_string.find(')')
        
        if opening_index != -1 and closing_index != -1:
            substring = input_string[opening_index + 1 : closing_index]
            return substring

    def add_list(st):
        with open(csv1_path, 'r') as file:
            reader = csv.reader(file)
            li = []
            for row in reader:
                r1=row[1].split(",")
                if findsubstring(st) != -1:
                    index = findsubstring(st)

                    li.append(print_string_between_parentheses(r1[index]))
                else:
                    li.append("Absent")
            return li

    def add_regID(filename):
        with open(filename, 'r') as file:
            reader = csv.reader(file)
            li = []
            listTRV = []
            for row in reader:
                if 'TRV' in row[0]:
                    listTRV.append(row[0])
                elif 'PKD' in row[0]:
                    listTRV.append(row[0])
            return listTRV


    grade_points = {'S': 10, 'A+': 9.0, 'A': 8.5, 'B+': 8.0, 'B': 7.5, 'C+': 7.0, 'C': 6.5, 'D': 6.0, 'P': 5.5, 'F': 0, 'FE': 0, 'Ab': 0}
    credits = {'CYT100': 4, 'EST100': 3, 'EST130': 4, 'CYL120': 1, 'ESL130': 1, 'MAT102': 4, 'HUN102': 0, 'EST102': 4}
    r1=[]
   
    time.sleep(1)
    listTRV = add_regID(csv1_path)
    l=list(credits.keys())
    listc1=add_list(l[0])
    listc2=add_list(l[1])
    listc3=add_list(l[2])
    listc4=add_list(l[3])
    listc5=add_list(l[4])
    listc6=add_list(l[5])
    listc7=add_list(l[6])
    listc8=add_list(l[7])
    with open(csv1_path, 'r') as csvfile:
        reader = csv.reader(csvfile)
        for row in reader:
            for item in row:
                if "TRV20IT001"==item :
                    h=next(reader)
                    listc8[0]=print_string_between_parentheses(h[1])
                elif  "TRV20IT002"==item:
                    h=next(reader)
                    listc8[1]=print_string_between_parentheses(h[1])                        
                       

    listnames=[listc1,listc2,listc3,listc4,listc5,listc6,listc7,listc8]
    # Calculate SGPA
                    
    sgpa=0
    grade_counts = {course_code: {'S': 0, 'A+': 0, 'A': 0, 'B+': 0, 'B': 0, 'C+': 0, 'C': 0, 'D': 0, 'P': 0, 'F': 0, 'FE': 0, 'Absent': 0} for course_code in credits}
    SGPA=[]
    grade_points = {'S': 10, 'A+': 9.0, 'A': 8.5, 'B+': 8.0, 'B': 7.5, 'C+': 7.0, 'C': 6.5, 'D': 6.0, 'P': 5.5, 'F': 0, 'FE': 0, 'Absent': 0}
    for i in range(length):
        has_failing_grade = False
        total_credit_points = 0
        total_credits = 0
        j=0
        for course_code in credits:
            # grade = eval("list" + course_code)[i]
        
            li=listnames[j]
            # print(li)
            grade=li[i]
            # print(grade,end=" ")
            if grade in grade_points:
                if grade in ['F', 'FE', 'Absent']:
                    grade_counts[course_code][grade] += 1
                    has_failing_grade = True
                else:
                    grade_counts[course_code][grade] += 1
                total_credit_points += credits[course_code] * grade_points[grade]
                total_credits += credits[course_code]
                j=j+1
        if has_failing_grade:
            sgpa = 0
        else:
            sgpa = total_credit_points / 21
        SGPA.append(sgpa)

    def find_top_students(filename, sgpa_list, reg_id_list):
        student_data = zip(reg_id_list, sgpa_list)
        sorted_students = sorted(student_data, key=lambda x: x[1], reverse=True)
        top_students = sorted_students[:3]
        return top_students
        

    # function to check if a particular combination of stu and sem already exists in the table
    # def check_combination_exists(stu_id, sem):
    #     combination_exists = Student.objects.filter(stu_id=stu_id, sem=sem).exists()
    #     return combination_exists
    # # to insert new rows to the table
    # def insert_student(stu_id, sem, sgpa):
    #                 student = Student(stu_id=stu_id, sem=sem, sgpa=sgpa)
    #                 student.save()
    # Write the result to a new CSV file
    time.sleep(1)
    output_file='output4.csv'
    create_empty_csv(uploads_dir,output_file)
    csv4_path = os.path.join(uploads_dir, "output4.csv")
    def update_student(stu_id, sem, sgpa):
     try:
        # Try to retrieve the existing student record
        student = Student.objects.get(stu_id=stu_id, sem=sem)
        
        # If a matching record is found, update the existing row
        student.sgpa = sgpa
        student.save()
     except Student.DoesNotExist:
        # If the combination doesn't exist, create a new row
        student = Student(stu_id=stu_id, sem=sem, sgpa=sgpa)
        student.save()
    with open(csv4_path, 'w', newline='') as file1:
        writer = csv.writer(file1)
        writer.writerow(['RegID', l[0], l[1],l[2],l[3],l[4],l[5],l[6],l[7],'SGPA'])
        for i in range(length):
            
             update_student(listTRV[i],'S2',SGPA[i])


            #save data add id,add sem,add sgpa
            #get all sgpas of listTRV[i]
            #cgpa calculation
            
             writer.writerow([listTRV[i],add_list(l[0])[i],add_list(l[1])[i],add_list(l[2])[i],add_list(l[3])[i],add_list(l[4])[i],
                            add_list(l[5])[i],add_list(l[6])[i],listc8[i],SGPA[i]])#write cgpa

# save in this loop^^(database)

    top_students=find_top_students(csv1_path, SGPA, listTRV)
    # Convert the CSV to Excel format
    wb = Workbook()
    ws = wb.active

    time.sleep(1)
    with open(csv4_path, 'r') as file2:
        reader = csv.reader(file2)
        for row in reader:
            ws.append(row)
    ws.append([])  # Add an empty row for separation
    ws.append(["Top Three Students"])
    ws.append(["RegID", "SGPA"])
    for student in top_students:
        ws.append(student)
    # Add grade counts per subject to the Excel file
    ws.append([])  # Add an empty row for separation
    ws.append(["Grade Counts per Subject"])
    for course_code, grade_count in grade_counts.items():
        ws.append([])  # Add an empty row for separation
        ws.append(["Subject: " + course_code])
        ws.append(["Grade", "Count"])
        for grade, count in grade_count.items():
            ws.append([grade, count])
    ws.append([])  # Add an empty row for separation
    ws.append(["Supplementary Count for Each Course"])
    ws.append(['Course', 'Absent', 'F', 'FE'])
    # for course_code in dict:
    #     ws.append([course_code, (grade_count.get(course_code, 0)), count_F.get(course_code, 0), count_FE.get(course_code, 0)])
    for course_code,grades in grade_counts.items():
        fe_count=grades.get('FE',0)
        f_count=grades.get('F',0)
        ab_count=grades.get('Absent',0)
        ws.append([course_code,ab_count, f_count, fe_count])
       

    dict_passfail={course_code: {'passcount': 0, 'failcount': 0} for course_code in credits}
    for i in range(length):
        j=0
        for course_code in credits:
            # grade = eval("list" + course_code)[i]
        
            li=listnames[j]
            # print(li)
            grade=li[i]
            # print(grade,end=" ")
            if grade in grade_points:
                if grade in ['F', 'FE', 'Absent']:
                    dict_passfail[course_code]['failcount'] += 1
                else:
                    dict_passfail[course_code]['passcount'] += 1
                j=j+1

    dict_passpercent={course_code: 0 for course_code in credits}
    percent=0
    list_codes=list(credits.keys())
    j=0
    for i in dict_passfail.values():
            course_code=list_codes[j]
            # print(i['passcount'])
            percent=i['passcount']*100/(i['passcount']+i['failcount'])
            dict_passpercent[course_code]=percent
            j=j+1
            
            # print(percent)

    # print(dict_passfail)
    ws.append([])  # Add an empty row for separation
    ws.append(["Pass Percentage for Each Course"])
    ws.append(['Course','Percent'])
    # for course_code in dict:
    #     ws.append([course_code, (grade_count.get(course_code, 0)), count_F.get(course_code, 0), count_FE.get(course_code, 0)])
    for course_code,percent in dict_passpercent.items():
            ws.append([course_code,percent])
            plt.bar(course_code,percent)
            plt.xlabel('Course')
            plt.ylabel('Passpercent')
            plt.title('Bar Chart')

# Save the figure
            plt.savefig('chart.png')
            img = Image('chart.png')
            img.width = 400
            img.height = 300
            ws.add_image(img, 'E205')
    CGPA=[]
    cgpa=0
    for i in range(length):
     student1=Student.objects.get(stu_id=listTRV[i], sem='S1')
     student = Student.objects.get(stu_id=listTRV[i], sem='S2')
     if student1.sgpa==0 or student.sgpa==0:
        cgpa=0
     else:
        cgpa=((student1.sgpa*17)+(student.sgpa*21))/38
     CGPA.append(cgpa)
    ws.append([])
    ws.append(["CGPA Calculation"])
    ws.append(['RegID','CGPA'])
    for i in range(length):
        ws.append([listTRV[i],CGPA[i]])
    output_path = os.path.join("media", "converted", "output4.xlsx")
    # workbook = Workbook()
    wb.save(output_path)

    pythoncom.CoInitialize()
    # Open the Excel application
    excel = win32.gencache.EnsureDispatch('Excel.Application')
# Open the workbook
    wb = excel.Workbooks.Open(r'C:\Users\Reon Augustine\Desktop\MiniProject\virtualresult\result_analysis\media\converted\output4.xlsx')

    # Save the workbook as a PDF file
    pdf_file_path = r'C:\Users\Reon Augustine\Desktop\MiniProject\virtualresult\result_analysis\media\converted\output4.pdf'
    xlFixedFormat = 0
    try:
        wb.ExportAsFixedFormat(xlFixedFormat, pdf_file_path)
        print(f"PDF file saved successfully: {pdf_file_path}")
    except Exception as e:
        print(f"Error occurred while saving PDF: {e}")

    # Close the workbook
    wb.Close()

    # Quit Excel application
    excel.Quit()

    return output_path




def analyze_s5(request):
    
    # Get the current directory
    current_dir = os.getcwd()

    # Define the path to the media directory
    media_dir = os.path.join(current_dir, 'media')

    # Define the path to the uploads directory inside the media directory
    uploads_dir = os.path.join(media_dir, 'uploads')

    # Create the uploads directory if it doesn't exist
    if not os.path.exists(uploads_dir):
        os.makedirs(uploads_dir)

    # Initialize csv_path variable
    csv_path = None

    try:
        file = tabula.read_pdf(uploaded_file, pages="all")
        tabula.convert_into(uploaded_file, os.path.join(uploads_dir, "output2.csv"), output_format="csv", pages="all")
        csv_path = os.path.join(uploads_dir, "output2.csv")
    except Exception as e:
        print("Error:", e)


    file = file

    li=[]
    
    # Open the CSV file
    with open(csv_path, 'r') as file:
        # Create a CSV reader object
        reader = csv.reader(file)

        # Iterate over each row in the CSV file
        for row in reader:
            # Access the values in each row
            # Example: Print the values in each row
            for i in row:
                if i.startswith('TRV20IT'):
                    li.append((row))
                elif i.startswith('LTRV20IT'):
                    li.append((row))
                elif i.startswith('PKD20IT'):
                    li.append((row))
                else:
                    continue
    time.sleep(1)
    output_file='output.csv'
    create_empty_csv(uploads_dir,output_file)
    length=len(li)
    csv1_path = os.path.join(uploads_dir, "output.csv")
    with open(csv1_path, 'w', newline='') as file1:
        writer = csv.writer(file1)
        for i in range(length):
            writer.writerow(li[i])


    # Function to find index of a substring from a list
    def findsubstring(s1):
        with open(csv1_path, 'r') as file:
            reader=csv.reader(file)
            for s in reader:
                r=s[1].split(",")
                for i, string in enumerate(r):
                    if s1 in string:
                        return i
                return -1 
                
    def print_string_between_parentheses(input_string):
        opening_index = input_string.find('(')
        closing_index = input_string.find(')')
        
        if opening_index != -1 and closing_index != -1:
            substring = input_string[opening_index + 1 : closing_index]
            return substring

    def add_list(st):
        with open(csv1_path, 'r') as file:
            reader = csv.reader(file)
            li = []
            for row in reader:
                r1=row[1].split(",")
                if findsubstring(st) != -1:
                    index = findsubstring(st)

                    li.append(print_string_between_parentheses(r1[index]))
                else:
                    li.append("Absent")
            return li

    def add_regID(filename):
        with open(filename, 'r') as file:
            reader = csv.reader(file)
            li = []
            listTRV = []
            for row in reader:
                if 'TRV' in row[0]:
                    listTRV.append(row[0])
                elif 'PKD' in row[0]:
                    listTRV.append(row[0])
            return listTRV


    grade_points = {'S': 10, 'A+': 9.0, 'A': 8.5, 'B+': 8.0, 'B': 7.5, 'C+': 7.0, 'C': 6.5, 'D': 6.0, 'P': 5.5, 'F': 0, 'FE': 0, 'Ab': 0}
    dict={}
    credits = {'MAT101':4,'PHT100':4,'EST110':3,'EST120':4,'HUN101':0,'PHL120':1,'ESL120':1,'CYT100': 4, 'EST100': 4, 'EST130': 4, 'CYL120': 1, 'ESL130': 1, 'MAT102': 4, 'HUN102': 0, 'EST102': 4,'MAT203':4,'ITT201':4,'ITT203':4,'ITT205':4,'EST200':2,'MCN201':0,'ITL201':2,'ITL203':2,'MAT208':4,'ITT202':4,'ITT204':4,'ITT206':4,'HUT200':2,'MCN202':0,'ITL202':2,'ITL204':2,'ITT301':4,'ITT303':4,'ITT305':4,'ITT307':4,'ITT309':3,'MCN301':0,'ITL331':2,'ITL333':2,'ITT302':4,'ITT304':4,'ITT306':4,'ITT308':1,'ITL332':2,'ITD334':2,'ITT312':3,'ITT332':3}
    r1=[]
    # def get_first_row_as_dict(file_path):
    time.sleep(1)
    with open(csv1_path, 'r') as csvfile:
        reader = csv.reader(csvfile)
        for i in reader:
            r1=i[1].split(",")
            # print(r1)
        for j in r1:
            index=j.index("(")
            # print(index)
            dict[j[index-6:index]]=credits[j[index-6:index]]

    for i in dict:
        list_name="list"+str(i)
        list_name=add_list(str(i))
        # print(list_name)
        length=len(list_name)

    listTRV = add_regID(csv1_path)
    l=list(dict.keys())
    # with open('output4.csv', 'w', newline='') as file1:
    #     writer = csv.writer(file1)
        
    #     writer.writerow(['RegID', l[0], l[1],l[2],l[3],l[4],l[5],l[6],l[7]])
    #     for i in range(length):
    #         writer.writerow([listTRV[i],add_list(l[0])[i],add_list(l[1])[i],add_list(l[2])[i],add_list(l[3])[i],add_list(l[4])[i],
    #                          add_list(l[5])[i],add_list(l[6])[i],add_list(l[7])[i]])
    listc1=add_list(l[0])
    listc2=add_list(l[1])
    listc3=add_list(l[2])
    listc4=add_list(l[3])
    listc5=add_list(l[4])
    listc6=add_list(l[5])
    listc7=add_list(l[6])
    listc8=add_list(l[7])

    listnames=[listc1,listc2,listc3,listc4,listc5,listc6,listc7,listc8]
    # Calculate SGPA
                    
    sgpa=0
    grade_counts = {course_code: {'S': 0, 'A+': 0, 'A': 0, 'B+': 0, 'B': 0, 'C+': 0, 'C': 0, 'D': 0, 'P': 0, 'F': 0, 'FE': 0, 'Absent': 0} for course_code in dict}
    SGPA=[]
    grade_points = {'S': 10, 'A+': 9.0, 'A': 8.5, 'B+': 8.0, 'B': 7.5, 'C+': 7.0, 'C': 6.5, 'D': 6.0, 'P': 5.5, 'F': 0, 'FE': 0, 'Absent': 0}
    credits = {'MAT101':4,'PHT101':4,'EST110':3,'EST120':4,'HUN101':0,'PHL120':1,'ESL120':1,'CYT100': 4, 'EST100': 4, 'EST130': 4, 'CYL120': 1, 'ESL130': 1, 'MAT102': 4, 'HUN102': 0, 'EST102': 4,'MAT203':4,'ITT201':4,'ITT203':4,'ITT205':4,'EST200':2,'MCN201':0,'ITL201':2,'ITL203':2,'MAT208':4,'ITT202':4,'ITT204':4,'ITT206':4,'HUT200':2,'MCN202':0,'ITL202':2,'ITL204':2,'ITT301':4,'ITT303':4,'ITT305':4,'ITT307':4,'ITT309':3,'MCN301':0,'ITL331':2,'ITL333':2,'ITT302':4,'ITT304':4,'ITT306':4,'ITT308':1,'ITL332':2,'ITD334':2,'ITT312':3,'ITT332':3}
    for i in range(length):
        has_failing_grade = False
        total_credit_points = 0
        total_credits = 0
        j=0
        for course_code in dict:
            # grade = eval("list" + course_code)[i]
        
            li=listnames[j]
            # print(li)
            grade=li[i]
            # print(grade,end=" ")
            if grade in grade_points:
                if grade in ['F', 'FE', 'Absent']:
                    grade_counts[course_code][grade] += 1
                    has_failing_grade = True
                else:
                    grade_counts[course_code][grade] += 1
                total_credit_points += credits[course_code] * grade_points[grade]
                total_credits += credits[course_code]
                j=j+1
        if has_failing_grade:
            sgpa = 0
        else:
            sgpa = total_credit_points / total_credits
        SGPA.append(sgpa)
    SGPA1=SGPA

    def find_top_students(filename, sgpa_list, reg_id_list):
        student_data = zip(reg_id_list, sgpa_list)
        sorted_students = sorted(student_data, key=lambda x: x[1], reverse=True)
        top_students = sorted_students[:3]
        return top_students
    def update_student(stu_id, sem, sgpa):
     try:
        # Try to retrieve the existing student record
        student = Student.objects.get(stu_id=stu_id, sem=sem)
        
        # If a matching record is found, update the existing row
        student.sgpa = sgpa
        student.save()
     except Student.DoesNotExist:
        # If the combination doesn't exist, create a new row
        student = Student(stu_id=stu_id, sem=sem, sgpa=sgpa)
        student.save()
    

        
    # Write the result to a new CSV file
    time.sleep(1)
    output_file='output4.csv'
    create_empty_csv(uploads_dir,output_file)
    csv4_path = os.path.join(uploads_dir, "output4.csv")
    with open(csv4_path, 'w', newline='') as file1:
        writer = csv.writer(file1)
        writer.writerow(['RegID', l[0], l[1],l[2],l[3],l[4],l[5],l[6],l[7],'SGPA'])
        for i in range(length):
            update_student(listTRV[i],'S5',SGPA[i])
            writer.writerow([listTRV[i],add_list(l[0])[i],add_list(l[1])[i],add_list(l[2])[i],add_list(l[3])[i],add_list(l[4])[i],
                            add_list(l[5])[i],add_list(l[6])[i],add_list(l[7])[i],SGPA[i]])

    
     


    top_students=find_top_students(csv1_path, SGPA, listTRV)
    # Convert the CSV to Excel format
    wb = Workbook()
    ws = wb.active

    time.sleep(1)
    with open(csv4_path, 'r') as file2:
        reader = csv.reader(file2)
        for row in reader:
            ws.append(row)
    ws.append([])  # Add an empty row for separation
    ws.append(["Top Three Students"])
    ws.append(["RegID", "SGPA"])
    for student in top_students:
        ws.append(student)
    # Add grade counts per subject to the Excel file
    ws.append([])  # Add an empty row for separation
    ws.append(["Grade Counts per Subject"])
    for course_code, grade_count in grade_counts.items():
        ws.append([])  # Add an empty row for separation
        ws.append(["Subject: " + course_code])
        ws.append(["Grade", "Count"])
        for grade, count in grade_count.items():
            ws.append([grade, count])
    ws.append([])  # Add an empty row for separation
    ws.append(["Supplementary Count for Each Course"])
    ws.append(['Course', 'Absent', 'F', 'FE'])
    # for course_code in dict:
    #     ws.append([course_code, (grade_count.get(course_code, 0)), count_F.get(course_code, 0), count_FE.get(course_code, 0)])
    for course_code,grades in grade_counts.items():
        fe_count=grades.get('FE',0)
        f_count=grades.get('F',0)
        ab_count=grades.get('Absent',0)
        ws.append([course_code,ab_count, f_count, fe_count])

    dict_passfail={course_code: {'passcount': 0, 'failcount': 0} for course_code in dict}
    for i in range(length):
        j=0
        for course_code in dict:
            # grade = eval("list" + course_code)[i]
        
            li=listnames[j]
            # print(li)
            grade=li[i]
            # print(grade,end=" ")
            if grade in grade_points:
                if grade in ['F', 'FE', 'Absent']:
                    dict_passfail[course_code]['failcount'] += 1
                else:
                    dict_passfail[course_code]['passcount'] += 1
                j=j+1

    dict_passpercent={course_code: 0 for course_code in dict}
    percent=0
    list_codes=list(dict.keys())
    j=0
    for i in dict_passfail.values():
            course_code=list_codes[j]
            # print(i['passcount'])
            percent=i['passcount']*100/(i['passcount']+i['failcount'])
            dict_passpercent[course_code]=percent
            j=j+1
            
            # print(percent)

    # print(dict_passfail)
    ws.append([])  # Add an empty row for separation
    ws.append(["Pass Percentage for Each Course"])
    ws.append(['Course','Percent'])
    # for course_code in dict:
    #     ws.append([course_code, (grade_count.get(course_code, 0)), count_F.get(course_code, 0), count_FE.get(course_code, 0)])
    for course_code,percent in dict_passpercent.items():
            ws.append([course_code,percent])
            plt.bar(course_code,percent)
            plt.xlabel('Course')
            plt.ylabel('Passpercent')
            plt.title('Bar Chart')

# Save the figure
            plt.savefig('chart.png')
            img = Image('chart.png')
            img.width = 400
            img.height = 300
            ws.add_image(img, 'E213')

    CGPA=[]
    cgpa=0
    for i in range(length):
     try:
         student1=Student.objects.get(stu_id=listTRV[i], sem='S1')
         student2=Student.objects.get(stu_id=listTRV[i], sem='S2')
         student3=Student.objects.get(stu_id=listTRV[i], sem='S3')
         student4=Student.objects.get(stu_id=listTRV[i], sem='S4')
         student5=Student.objects.get(stu_id=listTRV[i], sem='S5')
         if student1.sgpa==0 or student2.sgpa==0 or student3.sgpa==0 or student4.sgpa==0 or student5.sgpa==0:
             cgpa=0
         else:
          cgpa=((student1.sgpa*17)+(student2.sgpa*21)+(student3.sgpa*22)+(student4.sgpa*22)+(student5.sgpa*23))/105
         CGPA.append(cgpa)
     except Student.DoesNotExist:
         student3=Student.objects.get(stu_id=listTRV[i], sem='S3')
         student4=Student.objects.get(stu_id=listTRV[i], sem='S4')
         student5=Student.objects.get(stu_id=listTRV[i], sem='S5')
         if student3.sgpa==0 or student4.sgpa==0 or student5.sgpa==0:
             cgpa=0
         else:
          cgpa=((student3.sgpa*22)+(student4.sgpa*22)+(student5.sgpa*23))/67
         CGPA.append(cgpa)
    ws.append([])
    ws.append(["CGPA Calculation"])
    ws.append(['RegID','CGPA'])
    for i in range(length):
        ws.append([listTRV[i],CGPA[i]])    # wb.save('output4.xlsx')


    output_path = os.path.join("media", "converted", "output4.xlsx")
    # workbook = Workbook()
    wb.save(output_path)

    pythoncom.CoInitialize()
    # Open the Excel application
    excel = win32.gencache.EnsureDispatch('Excel.Application')

    # Open the workbook
    wb = excel.Workbooks.Open(r'C:\Users\Reon Augustine\Desktop\MiniProject\virtualresult\result_analysis\media\converted\output4.xlsx')

    # Save the workbook as a PDF file
    pdf_file_path = r'C:\Users\Reon Augustine\Desktop\MiniProject\virtualresult\result_analysis\media\converted\output4.pdf'
    xlFixedFormat = 0

    try:
        wb.ExportAsFixedFormat(xlFixedFormat, pdf_file_path)
        print(f"PDF file saved successfully: {pdf_file_path}")
    except Exception as e:
        print(f"Error occurred while saving PDF: {e}")

    # Close the workbook
    wb.Close()

    # Quit Excel application
    excel.Quit()

    return output_path